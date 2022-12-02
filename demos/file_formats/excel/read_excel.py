from statistics import mean

from openpyxl import load_workbook


wb = load_workbook(filename='portuguese_vinhoverde_redwine_quality.xlsx')

column_descriptions_sheet = wb['column-descriptions']

for row in column_descriptions_sheet[2: column_descriptions_sheet.max_row]:
    print(row[0].value, row[1].value)

data_sheet = wb['data']

print(mean([row[10].value for row in data_sheet[2: data_sheet.max_row]]))
print(mean([row[11].value for row in data_sheet[2: data_sheet.max_row]]))
